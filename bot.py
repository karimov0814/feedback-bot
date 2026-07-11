"""
Taklif & Shikoyat Telegram Bot
Railway: Webhook + Flask + PostgreSQL (Supabase) + Google Sheets
Ikki tilli: O'zbek / Rus
"""

import json
import logging
import os
import re
import asyncio
import threading
import time as time_module
import base64
import urllib.request
from io import BytesIO
from datetime import datetime
from zoneinfo import ZoneInfo
import psycopg2
import psycopg2.extras
import gspread
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from google.oauth2.service_account import Credentials
from flask import Flask, request, jsonify, Response
from flask_cors import CORS
from telegram import Update, WebAppInfo, InlineKeyboardButton, InlineKeyboardMarkup, InputFile, MenuButtonWebApp
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, filters, ContextTypes
from apscheduler.schedulers.background import BackgroundScheduler

# =====================================================
BOT_TOKEN      = os.environ.get("BOT_TOKEN", "")
ADMINS = {
    7780854728: "superadmin",
    1488298476: "superadmin",
    555648201:  "moderator",
}
ADMIN_IDS     = list(ADMINS.keys())
ADMIN_CHAT_ID = ADMIN_IDS[0]
MINI_APP_URL   = "https://karimov0814.github.io/ST77WOK/"
PORT           = int(os.environ.get("PORT", 5000))
WEBHOOK_URL    = os.environ.get("RAILWAY_PUBLIC_DOMAIN", "")
DATABASE_URL   = os.environ.get("DATABASE_URL", "")
SHEET_ID       = "13Dy2zeKPn4dmEHLKsjvtTi5qj4_X8aEefhJYrOpV2wI"

# Har safar bot qayta ishga tushganda (har bir Railway deploy) yangi qiymat oladi.
# Bu Mini App havolasiga avtomatik qo'shiladi va Telegram WebView'ning eski
# (keshlangan) index.html'ni ko'rsatib qolish muammosini oldini oladi —
# xodimlar /start bosmasdan ham har doim ENG YANGI versiyani ko'radi.
APP_VERSION = str(int(time_module.time()))


def mini_app_url(extra_query: str = "") -> str:
    """Mini App havolasini kesh-busting versiya bilan yasaydi.
    extra_query masalan 'admin=1' yoki 'view=my' bo'lishi mumkin (boshida '?' yoki '&' kerak emas)."""
    sep = "&" if "?" in MINI_APP_URL else "?"
    url = f"{MINI_APP_URL}{sep}v={APP_VERSION}"
    if extra_query:
        url += f"&{extra_query}"
    return url

# ---------- TUG'ILGAN KUN BOT ----------
BIRTHDAY_CHANNEL_ID  = os.environ.get("BIRTHDAY_CHANNEL_ID", "")   # tug'ilgan kun postlari yuboriladigan kanal
BIRTHDAY_PHOTO_PATH  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "birthday.jpg")
BIRTHDAY_TZ          = ZoneInfo("Asia/Tashkent")
BIRTHDAY_PREVIEW_HOUR, BIRTHDAY_PREVIEW_MIN = 9, 30   # adminlarga oldindan xabar
BIRTHDAY_POST_HOUR,    BIRTHDAY_POST_MIN    = 10, 0   # kanalga avtomatik post

# Supabase/Postgres serverida vaqt har doim UTC bo'yicha saqlanadi (created_at ustunlari),
# lekin foydalanuvchilarga har doim Toshkent vaqtida ko'rsatilishi kerak. Bazadan qaytgan
# "naive" (vaqt zonasiz) timestamp'larni ekranga chiqarishdan oldin shu funksiya orqali o'tkazamiz.
def utc_to_local(dt):
    if dt is None:
        return None
    return dt.replace(tzinfo=ZoneInfo("UTC")).astimezone(BIRTHDAY_TZ)

def _build_google_creds():
    raw_key = os.environ.get("GOOGLE_PRIVATE_KEY", "")
    if "\\n" in raw_key:
        raw_key = raw_key.replace("\\n", "\n")
    return {
        "type": "service_account",
        "project_id": "feedback-bot-499705",
        "private_key_id": "4e34a0b830ead7e9f6995dd6efe1b6f9e12f8c51",
        "private_key": raw_key,
        "client_email": "feedback-bot@feedback-bot-499705.iam.gserviceaccount.com",
        "client_id": "103704228425417838636",
        "auth_uri": "https://accounts.google.com/o/oauth2/auth",
        "token_uri": "https://oauth2.googleapis.com/token",
        "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
        "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/feedback-bot%40feedback-bot-499705.iam.gserviceaccount.com",
        "universe_domain": "googleapis.com"
    }

GOOGLE_CREDS = _build_google_creds()
# =====================================================

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
logger = logging.getLogger(__name__)

flask_app = Flask(__name__)
CORS(flask_app, resources={r"/*": {"origins": "*"}}, supports_credentials=False)

@flask_app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,POST,OPTIONS')
    return response

@flask_app.route('/', defaults={'path': ''}, methods=['OPTIONS'])
@flask_app.route('/<path:path>', methods=['OPTIONS'])
def options_handler(path):
    return jsonify({"ok": True}), 200

ptb_app = None
loop = None

# notif_id -> { admin_id: message_id, ... } — "Hammadan o'chirish" uchun
SENT_NOTIFICATIONS = {}


# ==================== GOOGLE SHEETS ====================

SHEET_HEADERS = [
    "ID", "Tur / Тип", "Kimdan / От кого", "Filial / Филиал",
    "Matn (UZ) / Текст (UZ)", "Matn (RU) / Текст (RU)",
    "Yuboruvchi / Отправитель", "Telefon / Телефон", "Anonim / Анонимно",
    "Til / Язык", "Vaqt / Время", "Status", "Rasm / Фото"
]

def get_sheet():
    try:
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        creds = Credentials.from_service_account_info(GOOGLE_CREDS, scopes=scopes)
        client = gspread.authorize(creds)
        sheet = client.open_by_key(SHEET_ID).sheet1

        if sheet.row_count == 0 or sheet.cell(1, 1).value != "ID":
            sheet.insert_row(SHEET_HEADERS, index=1)
        elif sheet.cell(1, 3).value != "Kimdan / От кого":
            sheet.update('A1', [SHEET_HEADERS])

        return sheet
    except Exception as e:
        logger.error(f"get_sheet xatolik: {e}")
        return None

def append_to_sheet(msg):
    try:
        sheet = get_sheet()
        if not sheet:
            return

        msg_type = msg.get("type", "")
        lang = msg.get("lang", "uz")
        is_anon = msg.get("anon", False)
        user_type = msg.get("user_type", "employee")

        if msg_type == "taklif":
            type_label = "Taklif / Предложение"
        else:
            type_label = "Shikoyat / Жалоба"

        kimdan_label = "Mehmon / Гость" if user_type == "guest" else "Xodim / Сотрудник"
        anon_label = "Ha / Да" if is_anon else "Yo'q / Нет"
        lang_label = "O'zbek" if lang == "uz" else "Русский"

        status_raw = msg.get("status", "new")
        if status_raw == "new":
            status_label = "Yangi / Новое"
        elif status_raw in ("progress", "in_progress"):
            status_label = "Ko'rib chiqilmoqda / В обработке"
        else:
            status_label = "Yakunlandi / Завершено"

        sheet.append_row([
            msg.get("id", ""),
            type_label,
            kimdan_label,
            msg.get("filial", ""),
            msg.get("text_uz") or msg.get("text", ""),
            msg.get("text_ru") or msg.get("text", ""),
            msg.get("sender", ""),
            msg.get("phone", "") or "",
            anon_label,
            lang_label,
            msg.get("time", ""),
            status_label,
            msg.get("photo_url", "") or "",
        ])
        logger.info(f"Google Sheets ga yozildi (ikki tilda): {msg.get('id')}")
    except Exception as e:
        logger.error(f"append_to_sheet xatolik: {e}")


# ==================== DATABASE ====================

def get_conn():
    return psycopg2.connect(DATABASE_URL, sslmode='require')

def init_db():
    try:
        conn = get_conn()
        cur = conn.cursor()

        cur.execute("""
            CREATE TABLE IF NOT EXISTS messages (
                id TEXT PRIMARY KEY,
                type TEXT,
                filial TEXT,
                text TEXT,
                text_uz TEXT,
                text_ru TEXT,
                lang TEXT DEFAULT 'uz',
                anon BOOLEAN DEFAULT FALSE,
                time TEXT,
                sender TEXT,
                status TEXT DEFAULT 'new',
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)

        for col, col_type, default in [
            ("text_uz", "TEXT", "''"),
            ("text_ru", "TEXT", "''"),
            ("lang", "TEXT", "'uz'"),
            ("user_type", "TEXT", "'employee'"),
            ("phone", "TEXT", "''"),
            ("photo_file_id", "TEXT", "''"),
            ("photo_url", "TEXT", "''"),
            ("real_sender", "TEXT", "''"),
            ("username", "TEXT", "''"),
        ]:
            try:
                cur.execute(f"ALTER TABLE messages ADD COLUMN IF NOT EXISTS {col} {col_type} DEFAULT {default}")
            except Exception:
                pass

        try:
            cur.execute("ALTER TABLE messages ADD COLUMN IF NOT EXISTS user_id BIGINT")
        except Exception:
            pass

        cur.execute("""
            CREATE TABLE IF NOT EXISTS contacts (
                user_id BIGINT PRIMARY KEY,
                phone TEXT,
                first_name TEXT,
                last_name TEXT,
                updated_at TIMESTAMP DEFAULT NOW()
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS replies (
                id SERIAL PRIMARY KEY,
                message_id TEXT REFERENCES messages(id) ON DELETE CASCADE,
                admin_id BIGINT,
                admin_name TEXT,
                text TEXT,
                is_read BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_replies_message_id ON replies(message_id)")

        conn.commit()
        cur.close()
        conn.close()
        logger.info("✅ Database tayyor (ikki tilli, javoblar bilan)")
        init_chat_table()
        init_employees_table()
    except Exception as e:
        logger.error(f"init_db xatolik: {e}")

def load_messages():
    try:
        conn = get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT m.*,
                   COALESCE((SELECT COUNT(*) FROM replies r WHERE r.message_id = m.id), 0) AS reply_count
            FROM messages m
            ORDER BY m.created_at DESC LIMIT 500
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return [dict(r) for r in rows]
    except Exception as e:
        logger.error(f"load_messages xatolik: {e}")
        return []


# ==================== EXCEL EKSPORT ====================

def load_messages_for_export(date_from=None, date_to=None, msg_type=None):
    """Admin panel: tanlangan davr (va ixtiyoriy tur) bo'yicha BARCHA murojaatlarni qaytaradi (500 talik limit yo'q)."""
    try:
        conn = get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        query = """
            SELECT m.*,
                   COALESCE((SELECT COUNT(*) FROM replies r WHERE r.message_id = m.id), 0) AS reply_count
            FROM messages m
            WHERE 1=1
        """
        params = []
        if date_from:
            query += " AND m.created_at >= %s"
            params.append(date_from)
        if date_to:
            query += " AND m.created_at <= %s"
            params.append(date_to)
        if msg_type in ("taklif", "shikoyat"):
            query += " AND m.type = %s"
            params.append(msg_type)
        query += " ORDER BY m.created_at ASC"
        cur.execute(query, params)
        rows = [dict(r) for r in cur.fetchall()]
        cur.close()
        conn.close()
        return rows
    except Exception as e:
        logger.error(f"load_messages_for_export xatolik: {e}")
        return []


def build_messages_excel(rows):
    """Murojaatlar ro'yxatidan formatlangan .xlsx fayl (bytes) yasaydi."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Murojaatlar"

    headers = [
        "ID", "Sana", "Turi", "Holat", "Kimdan", "Filial",
        "Matn (UZ)", "Matn (RU)", "Anonim",
        "Ko'rsatilgan ism", "Haqiqiy ism", "Username", "Telefon", "Javoblar soni",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="EA580C")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    type_labels = {"taklif": "Taklif", "shikoyat": "Shikoyat"}
    status_labels = {"new": "Yangi", "progress": "Jarayonda", "in_progress": "Jarayonda", "done": "Yakunlandi"}
    kimdan_labels = {"employee": "Xodim", "guest": "Mehmon"}

    for r in rows:
        created = r.get("created_at")
        sana = utc_to_local(created).strftime("%d.%m.%Y %H:%M") if created else (r.get("time") or "")
        username = r.get("username") or ""
        row_vals = [
            r.get("id", ""),
            sana,
            type_labels.get(r.get("type"), r.get("type") or ""),
            status_labels.get(r.get("status"), r.get("status") or ""),
            kimdan_labels.get(r.get("user_type"), r.get("user_type") or ""),
            r.get("filial") or "",
            r.get("text_uz") or r.get("text") or "",
            r.get("text_ru") or "",
            "Ha" if r.get("anon") else "Yo'q",
            r.get("sender") or "",
            r.get("real_sender") or (r.get("sender") if not r.get("anon") else "") or "",
            ("@" + username) if username else "",
            r.get("phone") or "",
            r.get("reply_count") or 0,
        ]
        ws.append(row_vals)

    widths = [15, 16, 10, 12, 9, 22, 45, 45, 9, 22, 22, 16, 15, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def save_message(msg):
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO messages (id, type, filial, text, text_uz, text_ru, lang, anon, time, sender, status, user_id,
                                   user_type, phone, photo_file_id, photo_url, real_sender, username)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (id) DO NOTHING
        """, (
            msg['id'], msg['type'], msg['filial'],
            msg['text'],
            msg.get('text_uz', msg['text']),
            msg.get('text_ru', msg['text']),
            msg.get('lang', 'uz'),
            msg['anon'],
            msg['time'], msg['sender'],
            msg.get('status', 'new'),
            msg.get('user_id'),
            msg.get('user_type', 'employee'),
            msg.get('phone', '') or '',
            msg.get('photo_file_id', '') or '',
            msg.get('photo_url', '') or '',
            msg.get('real_sender', '') or '',
            msg.get('username', '') or '',
        ))
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        logger.error(f"save_message xatolik: {e}")

def save_contact(user_id, phone, first_name='', last_name=''):
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO contacts (user_id, phone, first_name, last_name, updated_at)
            VALUES (%s, %s, %s, %s, NOW())
            ON CONFLICT (user_id) DO UPDATE SET
                phone = EXCLUDED.phone,
                first_name = EXCLUDED.first_name,
                last_name = EXCLUDED.last_name,
                updated_at = NOW()
        """, (user_id, phone, first_name, last_name))
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        logger.error(f"save_contact xatolik: {e}")

def get_contact_phone(user_id):
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT phone FROM contacts WHERE user_id = %s", (user_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        return row[0] if row else ''
    except Exception as e:
        logger.error(f"get_contact_phone xatolik: {e}")
        return ''

def delete_message_db(msg_id):
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("DELETE FROM messages WHERE id = %s", (msg_id,))
        affected = cur.rowcount
        conn.commit()
        cur.close()
        conn.close()
        return affected > 0
    except Exception as e:
        logger.error(f"delete_message xatolik: {e}")
        return False

def update_status_db(msg_id, new_status):
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("UPDATE messages SET status = %s WHERE id = %s", (new_status, msg_id))
        affected = cur.rowcount
        conn.commit()
        cur.close()
        conn.close()
        return affected > 0
    except Exception as e:
        logger.error(f"update_status xatolik: {e}")
        return False


# ==================== JAVOBLAR (REPLIES) ====================

def save_reply(message_id, admin_id, admin_name, text):
    try:
        conn = get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            INSERT INTO replies (message_id, admin_id, admin_name, text)
            VALUES (%s, %s, %s, %s)
            RETURNING id, message_id, admin_id, admin_name, text, created_at, is_read
        """, (message_id, admin_id, admin_name, text))
        row = cur.fetchone()
        conn.commit()
        cur.close()
        conn.close()
        return dict(row) if row else None
    except Exception as e:
        logger.error(f"save_reply xatolik: {e}")
        return None

def get_message_owner(message_id):
    try:
        conn = get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT id, user_id, type, filial FROM messages WHERE id = %s", (message_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        return dict(row) if row else None
    except Exception as e:
        logger.error(f"get_message_owner xatolik: {e}")
        return None

def load_my_messages(user_id):
    try:
        conn = get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM messages WHERE user_id = %s ORDER BY created_at DESC LIMIT 200", (user_id,))
        msgs = [dict(r) for r in cur.fetchall()]

        if msgs:
            ids = [m['id'] for m in msgs]
            cur.execute("SELECT * FROM replies WHERE message_id = ANY(%s) ORDER BY created_at ASC", (ids,))
            replies = [dict(r) for r in cur.fetchall()]
        else:
            replies = []

        cur.close()
        conn.close()

        by_msg = {}
        for r in replies:
            by_msg.setdefault(r['message_id'], []).append(r)

        for m in msgs:
            m_replies = by_msg.get(m['id'], [])
            m['replies'] = m_replies
            m['reply_count'] = len(m_replies)
            m['unread_count'] = sum(1 for r in m_replies if not r['is_read'])

        return msgs
    except Exception as e:
        logger.error(f"load_my_messages xatolik: {e}")
        return []

def mark_replies_read(message_id):
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("UPDATE replies SET is_read = TRUE WHERE message_id = %s AND is_read = FALSE", (message_id,))
        affected = cur.rowcount
        conn.commit()
        cur.close()
        conn.close()
        return affected
    except Exception as e:
        logger.error(f"mark_replies_read xatolik: {e}")
        return 0


# ==================== CHAT ====================

def init_chat_table():
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS chats (
                id SERIAL PRIMARY KEY,
                message_id TEXT REFERENCES messages(id) ON DELETE CASCADE,
                sender_type TEXT NOT NULL,  -- 'admin' or 'employee'
                sender_id BIGINT,
                sender_name TEXT,
                text TEXT NOT NULL,
                is_read BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_chats_message_id ON chats(message_id)")
        try:
            cur.execute("ALTER TABLE messages ADD COLUMN IF NOT EXISTS chat_closed BOOLEAN DEFAULT FALSE")
        except Exception:
            pass
        conn.commit()
        cur.close()
        conn.close()
        logger.info("✅ Chat jadvali tayyor")
    except Exception as e:
        logger.error(f"init_chat_table xatolik: {e}")

def get_chat_messages(message_id):
    try:
        conn = get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT id, message_id, sender_type, sender_id, sender_name, text, is_read,
                   TO_CHAR(created_at AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Tashkent', 'DD.MM.YYYY HH24:MI') as created_at
            FROM chats WHERE message_id = %s ORDER BY created_at ASC
        """, (message_id,))
        rows = [dict(r) for r in cur.fetchall()]
        # Also get chat_closed status
        cur.execute("SELECT chat_closed FROM messages WHERE id = %s", (message_id,))
        msg = cur.fetchone()
        cur.close()
        conn.close()
        return rows, (msg['chat_closed'] if msg else False)
    except Exception as e:
        logger.error(f"get_chat_messages xatolik: {e}")
        return [], False

def save_chat_message(message_id, sender_type, sender_id, sender_name, text):
    try:
        conn = get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            INSERT INTO chats (message_id, sender_type, sender_id, sender_name, text)
            VALUES (%s, %s, %s, %s, %s)
            RETURNING id, message_id, sender_type, sender_id, sender_name, text, created_at
        """, (message_id, sender_type, sender_id, sender_name, text))
        row = cur.fetchone()
        conn.commit()
        cur.close()
        conn.close()
        return dict(row) if row else None
    except Exception as e:
        logger.error(f"save_chat_message xatolik: {e}")
        return None

def set_chat_closed(message_id, closed):
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("UPDATE messages SET chat_closed = %s WHERE id = %s", (closed, message_id))
        conn.commit()
        cur.close()
        conn.close()
        return True
    except Exception as e:
        logger.error(f"set_chat_closed xatolik: {e}")
        return False

def get_chat_unread_count(message_id, for_admin=False):
    """Count unread messages for the OTHER side"""
    try:
        conn = get_conn()
        cur = conn.cursor()
        sender_type = 'employee' if for_admin else 'admin'
        cur.execute(
            "SELECT COUNT(*) FROM chats WHERE message_id = %s AND sender_type = %s AND is_read = FALSE",
            (message_id, sender_type)
        )
        count = cur.fetchone()[0]
        cur.close()
        conn.close()
        return count
    except Exception as e:
        logger.error(f"get_chat_unread_count xatolik: {e}")
        return 0

def mark_chat_read(message_id, reader_type):
    """Mark messages sent by OTHER side as read"""
    try:
        conn = get_conn()
        cur = conn.cursor()
        other = 'employee' if reader_type == 'admin' else 'admin'
        cur.execute(
            "UPDATE chats SET is_read = TRUE WHERE message_id = %s AND sender_type = %s AND is_read = FALSE",
            (message_id, other)
        )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        logger.error(f"mark_chat_read xatolik: {e}")


# ==================== BOT HANDLERS ====================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    chat_id = update.effective_chat.id
    is_admin = (chat_id in ADMIN_IDS)

    if is_admin:
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("📊 Admin Panel", web_app=WebAppInfo(url=mini_app_url("admin=1")))
        ]])
        await update.message.reply_text(
            "👋 Salom, Admin!\n\nXodimlarning murojaatlari shu yerga keladi.",
            reply_markup=keyboard
        )
    else:
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("💬 Taklif yoki Shikoyat yuborish", web_app=WebAppInfo(url=mini_app_url()))
        ]])
        await update.message.reply_text(
            f"👋 Salom, {user.first_name}!\n\n"
            f"💡 <b>Taklif</b> — g'oyalaringizni yuboring\n"
            f"⚠️ <b>Shikoyat</b> — muammolaringizni bildiring\n\n"
            f"🔒 Anonim yuborish imkoniyati mavjud.\n\n"
            f"👋 Привет, {user.first_name}!\n\n"
            f"💡 <b>Предложение</b> — поделитесь идеей\n"
            f"⚠️ <b>Жалоба</b> — сообщите о проблеме\n\n"
            f"🔒 Доступна анонимная отправка.\n\n"
            f"Boshlash / Начать 👇",
            reply_markup=keyboard,
            parse_mode="HTML"
        )

async def web_app_data(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Eski (sendData) yo'l — endi asosiy oqim /send orqali ketadi, lekin fallback sifatida qoldiramiz."""
    try:
        raw = update.message.web_app_data.data
        data = json.loads(raw)
        msg_type = data.get('type', '')
        filial   = data.get('filial', '')

        type_label_uz = "SHIKOYAT" if msg_type == "shikoyat" else "TAKLIF"
        emoji = "⚠️" if msg_type == "shikoyat" else "💡"
        notif = f"{emoji} Yangi {type_label_uz} — {filial} filialidan"

        admin_panel_kb = InlineKeyboardMarkup([[
            InlineKeyboardButton("📊 Admin Panel", web_app=WebAppInfo(url=mini_app_url("admin=1")))
        ]])

        for _aid in ADMIN_IDS:
            await context.bot.send_message(chat_id=_aid, text=notif, reply_markup=admin_panel_kb)
        await update.message.reply_text("✅ Murojaatingiz yuborildi! / Ваше обращение отправлено!\n\nRahmat / Спасибо!")

    except Exception as e:
        logger.error(f"web_app_data xatolik: {e}")
        await update.message.reply_text("❌ Xatolik yuz berdi. / Произошла ошибка.")


async def contact_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Mini App'dagi 'Raqamni ulashish' tugmasi bosilganda Telegram avtomatik shu yerga contact yuboradi."""
    try:
        contact = update.message.contact
        if not contact:
            return
        user_id = update.effective_user.id
        # Foydalanuvchi faqat OʻZINING raqamini ulashishi kerak
        if contact.user_id and contact.user_id != user_id:
            return
        save_contact(user_id, contact.phone_number, contact.first_name or '', contact.last_name or '')
        await update.message.reply_text(
            "✅ Telefon raqamingiz uchun rahmat! Endi murojaatingizni Mini App orqali yuborishingiz mumkin.\n\n"
            "✅ Спасибо за номер телефона! Теперь вы можете отправить обращение через Mini App."
        )
    except Exception as e:
        logger.error(f"contact_received xatolik: {e}")


async def delete_all_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """'🗑 Hammadan o'chirish' tugmasi — faqat superadmin uchun ishlaydi."""
    query = update.callback_query
    user_id = query.from_user.id

    if ADMINS.get(user_id) != "superadmin":
        await query.answer("Sizda bu amal uchun ruxsat yo'q", show_alert=True)
        return

    notif_id = query.data.split(":", 1)[1]
    sent_ids = SENT_NOTIFICATIONS.get(notif_id)

    if not sent_ids:
        await query.answer("Xabar topilmadi yoki muddati o'tgan", show_alert=True)
        return

    for _aid, _mid in sent_ids.items():
        try:
            await context.bot.delete_message(chat_id=_aid, message_id=_mid)
        except Exception as e:
            logger.error(f"delete_all_callback: {_aid} uchun o'chirib bo'lmadi: {e}")

    SENT_NOTIFICATIONS.pop(notif_id, None)
    await query.answer("Hammadan o'chirildi ✅")


# ==================== FLASK ROUTES ====================

@flask_app.route("/", methods=["GET"])
def index():
    return "Bot ishlayapti! / Бот работает! ✅", 200

@flask_app.route("/role/<int:telegram_id>", methods=["GET"])
def get_role(telegram_id):
    role = ADMINS.get(telegram_id)
    if role:
        return jsonify({"ok": True, "role": role, "is_admin": True})
    return jsonify({"ok": True, "role": "user", "is_admin": False})

@flask_app.route("/contact/<int:user_id>", methods=["GET"])
def contact_check(user_id):
    phone = get_contact_phone(user_id)
    return jsonify({"ok": True, "phone": phone})

@flask_app.route("/photo/<message_id>", methods=["GET"])
def get_photo(message_id):
    """Rasmni Telegramdan olib, to'g'ridan-to'g'ri frontendga uzatadi.
    Bot tokeni hech qachon brauzerga chiqmaydi — hammasi server ichida bajariladi."""
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT photo_file_id FROM messages WHERE id = %s", (message_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        if not row or not row[0]:
            return jsonify({"ok": False, "error": "Rasm topilmadi"}), 404

        file_id = row[0]
        file_future = asyncio.run_coroutine_threadsafe(ptb_app.bot.get_file(file_id), loop)
        tg_file = file_future.result(timeout=15)
        file_url = f"https://api.telegram.org/file/bot{BOT_TOKEN}/{tg_file.file_path}"

        with urllib.request.urlopen(file_url, timeout=15) as resp:
            img_bytes = resp.read()

        return Response(img_bytes, mimetype="image/jpeg", headers={"Cache-Control": "private, max-age=3600"})
    except Exception as e:
        logger.error(f"/photo xatolik: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

@flask_app.route(f"/webhook/{BOT_TOKEN}", methods=["POST"])
def webhook():
    data = request.get_json(force=True)
    update = Update.de_json(data, ptb_app.bot)
    asyncio.run_coroutine_threadsafe(ptb_app.process_update(update), loop)
    return jsonify({"ok": True})

@flask_app.route("/send", methods=["POST", "OPTIONS"])
def send_message():
    if request.method == "OPTIONS":
        return jsonify({"ok": True}), 200
    try:
        data = request.get_json(force=True)

        msg_type    = data.get('type', '')
        filial      = data.get('filial', '')
        text        = data.get('text', '')
        text_uz     = data.get('text_uz', text)
        text_ru     = data.get('text_ru', text)
        lang        = data.get('lang', 'uz')
        anon        = data.get('anon', False)
        time_str    = data.get('time', '')
        sender_name = data.get('sender_name', "Noma'lum / Неизвестно")
        username    = data.get('username', '')
        tg_user_id  = data.get('user_id') or None
        user_type   = data.get('user_type') or 'employee'
        if user_type not in ('employee', 'guest'):
            user_type = 'employee'
        photo_b64   = data.get('photo_base64')

        uname     = f" @{username}" if username else ""
        real_name = sender_name + uname

        if anon:
            display_name = "Anonim"
        else:
            display_name = real_name

        # Mehmon bo'lsa — Telegram orqali ulashilgan telefon raqamini bazadan olamiz
        phone = ''
        if user_type == 'guest' and tg_user_id:
            phone = get_contact_phone(int(tg_user_id))

        # Rasm bo'lsa — base64 dan decode qilamiz
        photo_bytes = None
        if photo_b64:
            try:
                if ',' in photo_b64:
                    photo_b64 = photo_b64.split(',', 1)[1]
                photo_bytes = base64.b64decode(photo_b64)
            except Exception as e:
                logger.error(f"photo decode xatolik: {e}")
                photo_bytes = None

        # ===== QISQA BILDIRISHNOMA (Telegram chatga shu boriladi) =====
        # Format: "⚠️ Yangi SHIKOYAT — Anhor filialidan (Mehmon)"
        type_label_uz = "SHIKOYAT" if msg_type == "shikoyat" else "TAKLIF"
        emoji = "⚠️" if msg_type == "shikoyat" else "💡"
        kimdan_lbl = "Mehmon" if user_type == "guest" else "Xodim"
        notif_text = f"{emoji} Yangi {type_label_uz} — {filial} filialidan ({kimdan_lbl})"

        msg_id = str(int(time_module.time() * 1000))

        # Har bir admin uchun rolga mos klaviatura va yuborilgan message_id larni saqlash
        sent_ids = {}
        photo_file_id = None
        for _aid in ADMIN_IDS:
            buttons = [InlineKeyboardButton("📊 Admin Panel", web_app=WebAppInfo(url=mini_app_url("admin=1")))]
            if ADMINS.get(_aid) == "superadmin":
                buttons.append(InlineKeyboardButton("🗑 Hammadan o'chirish", callback_data=f"delall:{msg_id}"))
            kb = InlineKeyboardMarkup([buttons])

            try:
                if photo_bytes:
                    if photo_file_id:
                        photo_to_send = photo_file_id
                    else:
                        bio = BytesIO(photo_bytes)
                        bio.name = 'photo.jpg'
                        photo_to_send = bio
                    future = asyncio.run_coroutine_threadsafe(
                        ptb_app.bot.send_photo(chat_id=_aid, photo=photo_to_send, caption=notif_text, reply_markup=kb),
                        loop
                    )
                    sent_msg = future.result(timeout=20)
                    if not photo_file_id:
                        photo_file_id = sent_msg.photo[-1].file_id
                else:
                    future = asyncio.run_coroutine_threadsafe(
                        ptb_app.bot.send_message(chat_id=_aid, text=notif_text, reply_markup=kb),
                        loop
                    )
                    sent_msg = future.result(timeout=10)
                sent_ids[_aid] = sent_msg.message_id
            except Exception as e:
                logger.error(f"admin bildirishnoma xatolik ({_aid}): {e}")

        SENT_NOTIFICATIONS[msg_id] = sent_ids

        # Rasm uchun havola — bot tokenini frontendga chiqarmaslik uchun
        # o'z serverimizdagi proxy manzilidan foydalanamiz (/photo/<id>)
        photo_url = ''
        if photo_file_id and WEBHOOK_URL:
            photo_url = f"https://{WEBHOOK_URL}/photo/{msg_id}"

        # To'liq ma'lumot — faqat DB ga (Mini App buni o'qiydi)
        msg = {
            "id":      msg_id,
            "type":    msg_type,
            "filial":  filial,
            "text":    text,
            "text_uz": text_uz,
            "text_ru": text_ru,
            "lang":    lang,
            "anon":    anon,
            "time":    time_str,
            "sender":  display_name,
            "status":  "new",
            "user_id": tg_user_id,
            "user_type": user_type,
            "phone": phone,
            "photo_file_id": photo_file_id or '',
            "photo_url": photo_url,
            # Anonim bo'lsa ham haqiqiy ism/username saqlanadi — faqat admin
            # Excel eksportida ko'rinadi, oddiy interfeys/chatda hech qachon chiqmaydi.
            "real_sender": sender_name,
            "username": username,
        }
        save_message(msg)

        # Google Sheets ga — har doim to'liq ism + ikki tilda matn
        sheet_msg = {
            "id":      msg_id,
            "type":    msg_type,
            "filial":  filial,
            "text":    text,
            "text_uz": text_uz,
            "text_ru": text_ru,
            "lang":    lang,
            "anon":    anon,
            "time":    time_str,
            "sender":  real_name,
            "status":  "new",
            "user_type": user_type,
            "phone": phone,
            "photo_url": photo_url,
        }
        threading.Thread(target=append_to_sheet, args=(sheet_msg,), daemon=True).start()

        logger.info(f"Yangi {msg_type} [{kimdan_lbl}]: {filial} — {'anonim' if anon else sender_name} [{lang}]")
        return jsonify({"ok": True})

    except Exception as e:
        logger.error(f"/send xatolik: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

@flask_app.route("/messages", methods=["GET", "OPTIONS"])
def get_messages():
    if request.method == "OPTIONS":
        return jsonify({"ok": True}), 200
    return jsonify({"ok": True, "messages": load_messages()})

@flask_app.route("/delete", methods=["POST", "OPTIONS"])
def delete_message():
    if request.method == "OPTIONS":
        return jsonify({"ok": True}), 200
    try:
        data   = request.get_json(force=True)
        msg_id = data.get('id')
        if not msg_id:
            return jsonify({"ok": False, "error": "id kerak"}), 400
        found = delete_message_db(msg_id)
        if not found:
            return jsonify({"ok": False, "error": "Xabar topilmadi"}), 404
        logger.info(f"Xabar o'chirildi: {msg_id}")
        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"/delete xatolik: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

@flask_app.route("/status", methods=["POST", "OPTIONS"])
def update_status():
    if request.method == "OPTIONS":
        return jsonify({"ok": True}), 200
    try:
        data       = request.get_json(force=True)
        msg_id     = data.get('id')
        new_status = data.get('status')

        if not msg_id or not new_status:
            return jsonify({"ok": False, "error": "id va status kerak"}), 400
        if new_status not in ["new", "progress", "in_progress", "done"]:
            return jsonify({"ok": False, "error": "Noto'g'ri status"}), 400

        found = update_status_db(msg_id, new_status)
        if not found:
            return jsonify({"ok": False, "error": "Xabar topilmadi"}), 404
        return jsonify({"ok": True})

    except Exception as e:
        logger.error(f"/status xatolik: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@flask_app.route("/my-messages/<int:user_id>", methods=["GET", "OPTIONS"])
def my_messages(user_id):
    if request.method == "OPTIONS":
        return jsonify({"ok": True}), 200
    return jsonify({"ok": True, "messages": load_my_messages(user_id)})


@flask_app.route("/reply", methods=["POST", "OPTIONS"])
def reply_message():
    if request.method == "OPTIONS":
        return jsonify({"ok": True}), 200
    try:
        data = request.get_json(force=True)
        message_id = data.get('message_id')
        admin_id   = data.get('admin_id')
        admin_name = (data.get('admin_name') or 'Admin').strip()
        text       = (data.get('text') or '').strip()

        if not message_id or not text:
            return jsonify({"ok": False, "error": "message_id va text kerak"}), 400
        if ADMINS.get(admin_id) is None:
            return jsonify({"ok": False, "error": "Sizda bu amal uchun ruxsat yo'q"}), 403

        reply = save_reply(message_id, admin_id, admin_name, text)
        if not reply:
            return jsonify({"ok": False, "error": "Javob saqlanmadi"}), 500

        owner = get_message_owner(message_id)
        if owner and owner.get('user_id'):
            try:
                type_label_uz = "SHIKOYAT" if owner.get('type') == 'shikoyat' else "TAKLIF"
                notif = (
                    f"✉️ Sizning {type_label_uz}ingizga javob keldi!\n\n"
                    f"💬 {text}\n\n"
                    f"Ko'rish uchun pastdagi tugmani bosing 👇"
                )
                kb = InlineKeyboardMarkup([[
                    InlineKeyboardButton("📂 Mening murojaatlarim", web_app=WebAppInfo(url=mini_app_url("view=my")))
                ]])
                future = asyncio.run_coroutine_threadsafe(
                    ptb_app.bot.send_message(chat_id=owner['user_id'], text=notif, reply_markup=kb),
                    loop
                )
                future.result(timeout=10)
            except Exception as e:
                logger.error(f"reply notif xatolik: {e}")

        logger.info(f"Javob yozildi: {message_id} <- admin {admin_id}")
        return jsonify({"ok": True, "reply": reply})

    except Exception as e:
        logger.error(f"/reply xatolik: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@flask_app.route("/reply/read", methods=["POST", "OPTIONS"])
def reply_read():
    if request.method == "OPTIONS":
        return jsonify({"ok": True}), 200
    try:
        data       = request.get_json(force=True)
        message_id = data.get('message_id')
        if not message_id:
            return jsonify({"ok": False, "error": "message_id kerak"}), 400
        mark_replies_read(message_id)
        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"/reply/read xatolik: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ==================== TUG'ILGAN KUNLAR (EMPLOYEES) ====================

def init_employees_table():
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS employees (
                id SERIAL PRIMARY KEY,
                fio TEXT NOT NULL,
                filial TEXT,
                birth_date DATE,
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        """)
        conn.commit()
        cur.close()
        conn.close()
        logger.info("✅ Employees va Settings jadvallari tayyor")
    except Exception as e:
        logger.error(f"init_employees_table xatolik: {e}")


def get_setting(key, default=None):
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT value FROM settings WHERE key = %s", (key,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        return row[0] if row else default
    except Exception as e:
        logger.error(f"get_setting xatolik: {e}")
        return default


def set_setting(key, value):
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO settings (key, value) VALUES (%s, %s)
            ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value
        """, (key, value))
        conn.commit()
        cur.close()
        conn.close()
        return True
    except Exception as e:
        logger.error(f"set_setting xatolik: {e}")
        return False


def get_birthday_post_time():
    """DB'da saqlangan vaqtni o'qiydi, bo'lmasa standart 10:00 ni qaytaradi. Natija: (soat, daqiqa)."""
    raw = get_setting("birthday_post_time", f"{BIRTHDAY_POST_HOUR:02d}:{BIRTHDAY_POST_MIN:02d}")
    try:
        h, m = raw.split(":")
        return int(h), int(m)
    except Exception:
        return BIRTHDAY_POST_HOUR, BIRTHDAY_POST_MIN


def replace_employees_from_excel(file_bytes):
    """Excel faylni o'qib, employees jadvalini to'liq yangilaydi (eskisini o'chirib, yangisini yozadi)."""
    df = pd.read_excel(BytesIO(file_bytes))
    total_rows = len(df)

    # Ustun nomlarini moslashtirish (bo'sh joy/registrdan mustaqil)
    col_map = {c.strip().lower(): c for c in df.columns}

    def find_col(*candidates):
        for cand in candidates:
            for lower, orig in col_map.items():
                if cand in lower:
                    return orig
        return None

    fio_col   = find_col("fio", "ф.и.о", "ism")
    filial_col = find_col("filial", "филиал")
    date_col  = find_col("tug'ilgan", "tugilgan", "sana", "рожден", "birth")

    if not fio_col or not date_col:
        raise ValueError("Excel faylda 'FIO' va 'Tug'ilgan_sana' ustunlari topilmadi")

    # dayfirst=True — O'zbekistonda sanalar KUN.OY.YIL formatida yoziladi (masalan 05.03.1998).
    # Buni belgilamasak, pandas ba'zi sanalarni OY.KUN.YIL deb noto'g'ri o'qiydi va
    # kun 12 dan katta bo'lgan sanalarni (13–31) xato deb tashlab yuboradi.
    df["_birth"] = pd.to_datetime(df[date_col], errors="coerce", dayfirst=True)

    missing_fio = df[fio_col].isna()
    missing_birth = df["_birth"].isna()
    skipped_count = int((missing_fio | missing_birth).sum())

    df_valid = df.dropna(subset=[fio_col, "_birth"])

    rows = []
    for _, r in df_valid.iterrows():
        fio = str(r[fio_col]).strip()
        filial = str(r[filial_col]).strip() if filial_col and pd.notna(r.get(filial_col)) else ""
        birth = r["_birth"].date()
        rows.append((fio, filial, birth))

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM employees")
    if rows:
        psycopg2.extras.execute_values(
            cur,
            "INSERT INTO employees (fio, filial, birth_date) VALUES %s",
            rows
        )
    conn.commit()
    cur.close()
    conn.close()
    return {"total": total_rows, "inserted": len(rows), "skipped": skipped_count}


def get_employees_summary():
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM employees")
        count = cur.fetchone()[0]
        cur.execute("SELECT MAX(created_at) FROM employees")
        last = cur.fetchone()[0]
        cur.close()
        conn.close()
        return {"count": count, "uploaded_at": utc_to_local(last).strftime("%d.%m.%Y %H:%M") if last else None}
    except Exception as e:
        logger.error(f"get_employees_summary xatolik: {e}")
        return {"count": 0, "uploaded_at": None}


def get_today_birthdays_db():
    try:
        today = datetime.now(BIRTHDAY_TZ)
        conn = get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT fio, filial FROM employees
            WHERE EXTRACT(DAY FROM birth_date) = %s AND EXTRACT(MONTH FROM birth_date) = %s
            ORDER BY fio
        """, (today.day, today.month))
        rows = [dict(r) for r in cur.fetchall()]
        cur.close()
        conn.close()
        return rows
    except Exception as e:
        logger.error(f"get_today_birthdays_db xatolik: {e}")
        return []


# ==================== KIRILL <-> LOTIN TRANSLITERATSIYA ====================
# Excel'dagi ism va filial ustunlari ba'zan lotin, ba'zan kirill alifbosida
# yozilishi mumkin. Kanalga post yuborilganda o'zbekcha qism har doim lotin,
# ruscha qism har doim kirill (yoki original ruscha so'z) bilan chiqishi uchun
# quyidagi transliteratsiya jadvali ishlatiladi. Bu — SO'ZMA-SO'Z TARJIMA emas,
# balki ISM kabi tarjima qilib bo'lmaydigan matnlar uchun yozuvni (skript)
# moslashtirish usuli.

_CYR_TO_LAT_MAP = {
    'А':'A','Б':'B','В':'V','Г':'G','Д':'D','Е':'Ye','Ё':'Yo','Ж':'J','З':'Z','И':'I',
    'Й':'Y','К':'K','Л':'L','М':'M','Н':'N','О':'O','П':'P','Р':'R','С':'S','Т':'T',
    'У':'U','Ф':'F','Х':'X','Ц':'Ts','Ч':'Ch','Ш':'Sh','Щ':'Sh','Ъ':'','Ы':'I','Ь':'',
    'Э':'E','Ю':'Yu','Я':'Ya','Ў':"O'",'Қ':'Q','Ғ':"G'",'Ҳ':'H',
    'а':'a','б':'b','в':'v','г':'g','д':'d','е':'ye','ё':'yo','ж':'j','з':'z','и':'i',
    'й':'y','к':'k','л':'l','м':'m','н':'n','о':'o','п':'p','р':'r','с':'s','т':'t',
    'у':'u','ф':'f','х':'x','ц':'ts','ч':'ch','ш':'sh','щ':'sh','ъ':'','ы':'i','ь':'',
    'э':'e','ю':'yu','я':'ya','ў':"o'",'қ':'q','ғ':"g'",'ҳ':'h',
}
_LAT_TO_CYR_DIGRAPHS = [
    ("O'", 'Ў'), ("o'", 'ў'), ("G'", 'Ғ'), ("g'", 'ғ'),
    ('SH', 'Ш'), ('Sh', 'Ш'), ('sh', 'ш'),
    ('CH', 'Ч'), ('Ch', 'Ч'), ('ch', 'ч'),
    ('YO', 'Ё'), ('Yo', 'Ё'), ('yo', 'ё'),
    ('YU', 'Ю'), ('Yu', 'Ю'), ('yu', 'ю'),
    ('YA', 'Я'), ('Ya', 'Я'), ('ya', 'я'),
]
_LAT_TO_CYR_SINGLE = {
    'A':'А','B':'Б','D':'Д','E':'Э','F':'Ф','G':'Г','H':'Ҳ','I':'И','J':'Ж','K':'К',
    'L':'Л','M':'М','N':'Н','O':'О','P':'П','Q':'Қ','R':'Р','S':'С','T':'Т','U':'У',
    'V':'В','X':'Х','Y':'Й','Z':'З',
    'a':'а','b':'б','d':'д','e':'э','f':'ф','g':'г','h':'ҳ','i':'и','j':'ж','k':'к',
    'l':'л','m':'м','n':'н','o':'о','p':'п','q':'қ','r':'р','s':'с','t':'т','u':'у',
    'v':'в','x':'х','y':'й','z':'з',
}
_CYRILLIC_RE = re.compile('[А-Яа-яЁёЎўҚқҒғҲҳ]')


def has_cyrillic(text):
    return bool(_CYRILLIC_RE.search(text or ""))


def cyrillic_to_latin(text):
    if not text:
        return text
    return "".join(_CYR_TO_LAT_MAP.get(ch, ch) for ch in text)


def latin_to_cyrillic(text):
    if not text:
        return text
    for lat, cyr in _LAT_TO_CYR_DIGRAPHS:
        text = text.replace(lat, cyr)
    return "".join(_LAT_TO_CYR_SINGLE.get(ch, ch) for ch in text)


def get_fio_bilingual(fio_raw):
    """Ism — tarjima qilinmaydi, faqat yozuvi (kirill/lotin) moslashtiriladi."""
    fio_raw = (fio_raw or "").strip()
    if has_cyrillic(fio_raw):
        return cyrillic_to_latin(fio_raw), fio_raw
    return fio_raw, latin_to_cyrillic(fio_raw)


# Ma'lum filiallar uchun HAQIQIY tarjima lug'ati (kod -> (o'zbekcha, ruscha)).
# Filial nomlari orasida haqiqiy ruscha so'zlar ham bor (masalan "Склад",
# "Руководство"), shuning uchun bularni oddiy transliteratsiya emas, balki
# to'g'ri tarjima orqali ko'rsatish kerak.
FILIAL_TRANSLATIONS = {
    "001": ("Bosh ofis", "Руководство"),
    "007": ("Xavfsizlik xizmati", "Служба безопасности"),
    "008": ("Call-Center", "Call-Center"),
    "010": ("Texnik bo'lim", "Тех отдел"),
    "011": ("Ombor RMS-1", "Склад РМС-1"),
    "012": ("Ombor RMS-2", "Склад РМС-2"),
    "013": ("Beruniy sexi", "Беруний цех"),
    "101": ("Glinka", "Глинка WOK"),
    "102": ("C1 Wok", "Ц1 WOK"),
    "103": ("Samarqand Darvoza", "Сам.дарвоза"),
    "104": ("C1 Street", "Ц1 Street"),
    "105": ("Scopus Mall", "Scopus Mall"),
    "106": ("Yunusobod Gallery", "Юнусабад Gallery"),
    "107": ("Beruniy", "Беруний"),
    "108": ("Eco", "Эко"),
    "109": ("Sergeli", "Сергели"),
    "110": ("Anhor", "Анхор"),
    "111": ("Novza", "Новза"),
    "112": ("Compass Mall", "Compass Mall"),
    "113": ("City Mall", "City Mall"),
    "114": ("Tuzel", "TUZEL WOK & STREET 77"),
    "115": ("Alfraganus", "Альфраганус"),
    "116": ("Chilonzor 20", "Чиланзар 20"),
    "601": ("Andijon Navruz", "Андижон_Навруз"),
}
_FILIAL_CODE_RE = re.compile(r'^\s*(\d+)\s*[.\-]?\s*')


def get_filial_bilingual(filial_raw):
    """Filial nomi — lug'atda bo'lsa haqiqiy tarjimasi, bo'lmasa yozuv moslashtirilgan holati qaytariladi."""
    filial_raw = (filial_raw or "").strip()
    if not filial_raw:
        return "", ""
    m = _FILIAL_CODE_RE.match(filial_raw)
    code = m.group(1).zfill(3) if m else None
    if code and code in FILIAL_TRANSLATIONS:
        return FILIAL_TRANSLATIONS[code]
    # Lug'atda yo'q (yangi filial) — eng yaxshi urinish: yozuvni moslashtiramiz
    if has_cyrillic(filial_raw):
        return cyrillic_to_latin(filial_raw), filial_raw
    return filial_raw, latin_to_cyrillic(filial_raw)


def build_birthday_post(birthdays):
    uz_lines, ru_lines = [], []
    for i, b in enumerate(birthdays, start=1):
        fio_uz, fio_ru = get_fio_bilingual(b.get('fio'))
        filial_uz, filial_ru = get_filial_bilingual(b.get('filial'))
        uz_lines.append(f"{i}) {fio_uz}" + (f" ({filial_uz})" if filial_uz else ""))
        ru_lines.append(f"{i}) {fio_ru}" + (f" ({filial_ru})" if filial_ru else ""))

    uz_text = (
        "🎊 Bugun jamoamizda kayfiyat ikki karra ko'tarinki!\n\n"
        + "\n".join(uz_lines) +
        "\n\n🎂 Tug'ilgan kuningiz muborak bo'lsin 🌟"
    )
    ru_text = (
        "🎊 Сегодня у нас в коллективе двойной праздник!\n\n"
        + "\n".join(ru_lines) +
        "\n\n🎂 Поздравляем с днём рождения 🌟"
    )
    return uz_text + "\n\n" + ("・" * 12) + "\n\n" + ru_text


def send_birthday_admin_preview():
    """09:30 — adminlarga bugungi tug'ilgan kunlar ro'yxatini yuboradi."""
    try:
        birthdays = get_today_birthdays_db()
        if birthdays:
            lines = "\n".join(f"• {b['fio']}" + (f" ({b['filial']})" if b.get("filial") else "") for b in birthdays)
            text = f"🔔 Eslatma: 30 daqiqadan so'ng kanalga tug'ilgan kun posti yuboriladi.\n\nBugun tug'ilgan kunlar:\n{lines}"
        else:
            text = "🔔 Bugun hech kimning tug'ilgan kuni yo'q. Kanalga post yuborilmaydi."

        for aid in ADMIN_IDS:
            try:
                future = asyncio.run_coroutine_threadsafe(
                    ptb_app.bot.send_message(chat_id=aid, text=text), loop
                )
                future.result(timeout=10)
            except Exception as e:
                logger.error(f"birthday preview -> admin {aid} xatolik: {e}")
    except Exception as e:
        logger.error(f"send_birthday_admin_preview xatolik: {e}")


def send_birthday_channel_post():
    """10:00 — kanalga tug'ilgan kun postini (UZ+RU, rasm bilan) yuboradi."""
    try:
        if not BIRTHDAY_CHANNEL_ID:
            logger.warning("BIRTHDAY_CHANNEL_ID sozlanmagan, post yuborilmadi")
            return
        birthdays = get_today_birthdays_db()
        if not birthdays:
            logger.info("Bugun tug'ilgan kun yo'q, kanalga post yuborilmadi")
            return

        caption = build_birthday_post(birthdays)

        async def _send():
            # Telegram caption limiti — 1024 belgi. Ro'yxat uzun bo'lsa,
            # rasmni captionsiz yuborib, matnni alohida xabar sifatida qo'shamiz.
            if len(caption) <= 1024:
                with open(BIRTHDAY_PHOTO_PATH, "rb") as photo:
                    await ptb_app.bot.send_photo(chat_id=BIRTHDAY_CHANNEL_ID, photo=photo, caption=caption)
            else:
                with open(BIRTHDAY_PHOTO_PATH, "rb") as photo:
                    await ptb_app.bot.send_photo(chat_id=BIRTHDAY_CHANNEL_ID, photo=photo)
                await ptb_app.bot.send_message(chat_id=BIRTHDAY_CHANNEL_ID, text=caption)

        future = asyncio.run_coroutine_threadsafe(_send(), loop)
        future.result(timeout=30)
        logger.info(f"🎂 Tug'ilgan kun posti kanalga yuborildi ({len(birthdays)} kishi)")
    except Exception as e:
        logger.error(f"send_birthday_channel_post xatolik: {e}")


birthday_scheduler = None

def refresh_menu_button():
    """Menu Button havolasidagi versiya raqamini yangilaydi — index.html GitHub'da
    yangilangan bo'lsa ham, botni qayta ishga tushirmasdan Telegram keshini avtomatik chetlab o'tadi."""
    global APP_VERSION
    APP_VERSION = str(int(time_module.time()))
    try:
        async def _set():
            await ptb_app.bot.set_chat_menu_button(
                menu_button=MenuButtonWebApp(text="Murojaatlar", web_app=WebAppInfo(url=mini_app_url()))
            )
        future = asyncio.run_coroutine_threadsafe(_set(), loop)
        future.result(timeout=15)
        logger.info(f"🔄 Menu Button versiyasi yangilandi (v={APP_VERSION})")
    except Exception as e:
        logger.error(f"refresh_menu_button xatolik: {e}")


def start_birthday_scheduler():
    global birthday_scheduler
    post_hour, post_min = get_birthday_post_time()
    prev_hour, prev_min = _minus_30_minutes(post_hour, post_min)

    birthday_scheduler = BackgroundScheduler(timezone=BIRTHDAY_TZ)
    birthday_scheduler.add_job(send_birthday_admin_preview, "cron",
                       hour=prev_hour, minute=prev_min, id="birthday_preview")
    birthday_scheduler.add_job(send_birthday_channel_post, "cron",
                       hour=post_hour, minute=post_min, id="birthday_post")
    # Har 15 daqiqada Mini App havolasini "yangi" qilib qo'yadi — shuning uchun
    # index.html'ni GitHub'da yangilaganingizda botni qayta ishga tushirish shart emas,
    # eng ko'pi bilan 15 daqiqadan keyin foydalanuvchilar yangi versiyani ko'radi.
    birthday_scheduler.add_job(refresh_menu_button, "interval", minutes=15, id="menu_button_refresh")
    birthday_scheduler.start()
    logger.info(f"✅ Tug'ilgan kun scheduler ishga tushdi ({prev_hour:02d}:{prev_min:02d} preview, {post_hour:02d}:{post_min:02d} post, Asia/Tashkent)")
    return birthday_scheduler


def _minus_30_minutes(hour, minute):
    total = hour * 60 + minute - 30
    if total < 0:
        total += 24 * 60
    return total // 60, total % 60


def reschedule_birthday_jobs(post_hour, post_min):
    """Admin panel orqali vaqt o'zgartirilganda ishlayotgan scheduler'ni qayta sozlaydi (restart shart emas)."""
    global birthday_scheduler
    prev_hour, prev_min = _minus_30_minutes(post_hour, post_min)
    if birthday_scheduler is None:
        return False
    birthday_scheduler.reschedule_job("birthday_preview", trigger="cron", hour=prev_hour, minute=prev_min)
    birthday_scheduler.reschedule_job("birthday_post", trigger="cron", hour=post_hour, minute=post_min)
    logger.info(f"🔄 Tug'ilgan kun vaqti yangilandi: {prev_hour:02d}:{prev_min:02d} preview, {post_hour:02d}:{post_min:02d} post")
    return True


# ==================== EMPLOYEES ROUTES ====================

@flask_app.route("/admin/employees", methods=["GET"])
def employees_summary():
    admin_id = request.args.get("admin_id", type=int)
    if not admin_id or ADMINS.get(admin_id) is None:
        return jsonify({"ok": False, "error": "Ruxsat yo'q"}), 403
    summary = get_employees_summary()
    summary["today_birthdays"] = get_today_birthdays_db()
    post_hour, post_min = get_birthday_post_time()
    summary["post_time"] = f"{post_hour:02d}:{post_min:02d}"
    return jsonify({"ok": True, **summary})


@flask_app.route("/admin/employees/upload", methods=["POST", "OPTIONS"])
def employees_upload():
    if request.method == "OPTIONS":
        return jsonify({"ok": True}), 200
    try:
        admin_id = request.form.get("admin_id", type=int)
        if not admin_id or ADMINS.get(admin_id) is None:
            return jsonify({"ok": False, "error": "Ruxsat yo'q"}), 403

        file = request.files.get("file")
        if not file or file.filename == "":
            return jsonify({"ok": False, "error": "Excel fayl biriktirilmagan"}), 400
        if not file.filename.lower().endswith((".xlsx", ".xls")):
            return jsonify({"ok": False, "error": "Faqat .xlsx yoki .xls fayl qabul qilinadi"}), 400

        stats = replace_employees_from_excel(file.read())
        logger.info(f"Employees excel yuklandi: {stats} (admin {admin_id})")
        return jsonify({"ok": True, "count": stats["inserted"], "total": stats["total"], "skipped": stats["skipped"]})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception as e:
        logger.error(f"/admin/employees/upload xatolik: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@flask_app.route("/admin/birthday-time", methods=["POST", "OPTIONS"])
def set_birthday_time():
    if request.method == "OPTIONS":
        return jsonify({"ok": True}), 200
    try:
        data = request.get_json(force=True)
        admin_id = data.get("admin_id")
        if not admin_id or ADMINS.get(int(admin_id)) is None:
            return jsonify({"ok": False, "error": "Ruxsat yo'q"}), 403

        time_str = (data.get("time") or "").strip()
        if not re.match(r"^([01]\d|2[0-3]):[0-5]\d$", time_str):
            return jsonify({"ok": False, "error": "Vaqt formati noto'g'ri (HH:MM bo'lishi kerak)"}), 400

        h, m = map(int, time_str.split(":"))
        set_setting("birthday_post_time", time_str)
        reschedule_birthday_jobs(h, m)
        logger.info(f"Tug'ilgan kun post vaqti o'zgartirildi: {time_str} (admin {admin_id})")
        return jsonify({"ok": True, "post_time": time_str})
    except Exception as e:
        logger.error(f"/admin/birthday-time xatolik: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ==================== EXPORT ROUTE ====================

@flask_app.route("/admin/export", methods=["GET"])
def export_messages():
    admin_id = request.args.get("admin_id", type=int)
    if not admin_id or ADMINS.get(admin_id) is None:
        return jsonify({"ok": False, "error": "Ruxsat yo'q"}), 403

    date_from_raw = request.args.get("date_from")  # 'YYYY-MM-DD'
    date_to_raw   = request.args.get("date_to")
    msg_type      = request.args.get("type")        # 'taklif' | 'shikoyat' | None

    date_from = f"{date_from_raw} 00:00:00" if date_from_raw else None
    date_to   = f"{date_to_raw} 23:59:59" if date_to_raw else None

    try:
        rows = load_messages_for_export(date_from, date_to, msg_type)
        excel_bytes = build_messages_excel(rows)
        filename = f"murojaatlar_{datetime.now(BIRTHDAY_TZ).strftime('%Y%m%d_%H%M')}.xlsx"
        logger.info(f"Excel eksport: {len(rows)} ta murojaat (admin {admin_id}, {date_from_raw}—{date_to_raw}, tur={msg_type})")
        return Response(
            excel_bytes,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        logger.error(f"/admin/export xatolik: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ==================== CHAT ROUTES ====================

@flask_app.route("/chat/messages/<message_id>", methods=["GET", "OPTIONS"])
def chat_get(message_id):
    if request.method == "OPTIONS":
        return jsonify({"ok": True}), 200
    reader = request.args.get('reader', 'employee')  # 'admin' or 'employee'
    mark_chat_read(message_id, reader)
    msgs, closed = get_chat_messages(message_id)
    return jsonify({"ok": True, "messages": msgs, "chat_closed": closed})


@flask_app.route("/chat/send", methods=["POST", "OPTIONS"])
def chat_send():
    if request.method == "OPTIONS":
        return jsonify({"ok": True}), 200
    try:
        data = request.get_json(force=True)
        message_id  = data.get('message_id')
        sender_type = data.get('sender_type', 'employee')  # 'admin' or 'employee'
        sender_id   = data.get('sender_id')
        sender_name = (data.get('sender_name') or '').strip() or ('Admin' if sender_type == 'admin' else 'Xodim')
        text        = (data.get('text') or '').strip()

        if not message_id or not text:
            return jsonify({"ok": False, "error": "message_id va text kerak"}), 400

        # Check if chat is closed
        _, closed = get_chat_messages.__wrapped__(message_id) if hasattr(get_chat_messages, '__wrapped__') else (None, None)
        # Direct check
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT chat_closed FROM messages WHERE id = %s", (message_id,))
            row = cur.fetchone()
            cur.close()
            conn.close()
            if row and row[0]:
                return jsonify({"ok": False, "error": "Chat yopilgan"}), 403
        except Exception:
            pass

        # Admin permission check
        if sender_type == 'admin' and sender_id and ADMINS.get(int(sender_id)) is None:
            return jsonify({"ok": False, "error": "Ruxsat yo'q"}), 403

        chat_msg = save_chat_message(message_id, sender_type, sender_id, sender_name, text)
        if not chat_msg:
            return jsonify({"ok": False, "error": "Xabar saqlanmadi"}), 500

        # Notify the other side via Telegram
        owner = get_message_owner(message_id)
        if sender_type == 'admin':
            # Notify employee
            if owner and owner.get('user_id'):
                try:
                    notif = (
                        f"💬 Sizga admin xabar yozdi!\n\n"
                        f"👮 {sender_name}: {text}\n\n"
                        f"Javob berish uchun 👇"
                    )
                    kb = InlineKeyboardMarkup([[
                        InlineKeyboardButton("💬 Chatni ochish", web_app=WebAppInfo(url=mini_app_url("view=my")))
                    ]])
                    future = asyncio.run_coroutine_threadsafe(
                        ptb_app.bot.send_message(chat_id=owner['user_id'], text=notif, reply_markup=kb),
                        loop
                    )
                    future.result(timeout=10)
                except Exception as e:
                    logger.error(f"chat notify employee xatolik: {e}")
        else:
            # Notify all admins
            if owner:
                type_label = "SHIKOYAT" if owner.get('type') == 'shikoyat' else "TAKLIF"
                notif = (
                    f"💬 {sender_name} chat orqali xabar yozdi!\n"
                    f"({type_label} — {owner.get('filial','')}) \n\n"
                    f"📝 {text}"
                )
                kb = InlineKeyboardMarkup([[
                    InlineKeyboardButton("📊 Admin Panel", web_app=WebAppInfo(url=mini_app_url("admin=1")))
                ]])
                for _aid in ADMIN_IDS:
                    try:
                        future = asyncio.run_coroutine_threadsafe(
                            ptb_app.bot.send_message(chat_id=_aid, text=notif, reply_markup=kb),
                            loop
                        )
                        future.result(timeout=10)
                    except Exception as e:
                        logger.error(f"chat notify admin {_aid} xatolik: {e}")

        return jsonify({"ok": True, "chat_message": chat_msg})
    except Exception as e:
        logger.error(f"/chat/send xatolik: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@flask_app.route("/chat/close", methods=["POST", "OPTIONS"])
def chat_close():
    if request.method == "OPTIONS":
        return jsonify({"ok": True}), 200
    try:
        data = request.get_json(force=True)
        message_id = data.get('message_id')
        admin_id   = data.get('admin_id')
        closed     = data.get('closed', True)

        if not message_id:
            return jsonify({"ok": False, "error": "message_id kerak"}), 400
        if not admin_id or ADMINS.get(int(admin_id)) is None:
            return jsonify({"ok": False, "error": "Faqat adminlar chatni yopa oladi"}), 403

        set_chat_closed(message_id, closed)

        # Notify employee
        owner = get_message_owner(message_id)
        if owner and owner.get('user_id'):
            try:
                if closed:
                    notif = "🔒 Sizning murojaatingiz bo'yicha chat yopildi.\n\nAdmin bilan muloqot yakunlandi."
                else:
                    notif = "🔓 Sizning murojaatingiz bo'yicha chat qayta ochildi."
                kb = InlineKeyboardMarkup([[
                    InlineKeyboardButton("📂 Mening murojaatlarim", web_app=WebAppInfo(url=mini_app_url("view=my")))
                ]])
                future = asyncio.run_coroutine_threadsafe(
                    ptb_app.bot.send_message(chat_id=owner['user_id'], text=notif, reply_markup=kb),
                    loop
                )
                future.result(timeout=10)
            except Exception as e:
                logger.error(f"chat close notify xatolik: {e}")

        action = "yopildi" if closed else "ochildi"
        logger.info(f"Chat {action}: {message_id} by admin {admin_id}")
        return jsonify({"ok": True, "chat_closed": closed})
    except Exception as e:
        logger.error(f"/chat/close xatolik: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@flask_app.route("/chat/unread/<int:user_id>", methods=["GET"])
def chat_unread_total(user_id):
    """Get total unread chat messages for an employee"""
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM chats c
            JOIN messages m ON c.message_id = m.id
            WHERE m.user_id = %s AND c.sender_type = 'admin' AND c.is_read = FALSE
        """, (user_id,))
        count = cur.fetchone()[0]
        cur.close()
        conn.close()
        return jsonify({"ok": True, "unread": count})
    except Exception as e:
        return jsonify({"ok": False, "unread": 0})


# ==================== MAIN ====================

async def setup_webhook():
    domain = WEBHOOK_URL
    if not domain:
        logger.warning("RAILWAY_PUBLIC_DOMAIN env yo'q!")
    else:
        webhook_address = f"https://{domain}/webhook/{BOT_TOKEN}"
        await ptb_app.bot.set_webhook(webhook_address)
        logger.info(f"Webhook o'rnatildi: {webhook_address}")

    # Doimiy "Murojaatlar" tugmasi — xabar yozish maydoni yonida doim turadi,
    # foydalanuvchi /start bosishi shart emas. Mini App'ning o'zi (checkRoleAndInit)
    # kim ekanini (admin/xodim) avtomatik aniqlaydi, shuning uchun hammaga bitta havola yetarli.
    try:
        await ptb_app.bot.set_chat_menu_button(
            menu_button=MenuButtonWebApp(text="Murojaatlar", web_app=WebAppInfo(url=mini_app_url()))
        )
        logger.info(f"✅ Doimiy Menu Button o'rnatildi (v={APP_VERSION})")
    except Exception as e:
        logger.error(f"Menu Button o'rnatishda xatolik: {e}")

def run():
    global ptb_app, loop

    init_db()

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)

    ptb_app = Application.builder().token(BOT_TOKEN).build()
    ptb_app.add_handler(CommandHandler("start", start))
    ptb_app.add_handler(MessageHandler(filters.StatusUpdate.WEB_APP_DATA, web_app_data))
    ptb_app.add_handler(MessageHandler(filters.CONTACT, contact_received))
    ptb_app.add_handler(CallbackQueryHandler(delete_all_callback, pattern=r"^delall:"))

    loop.run_until_complete(ptb_app.initialize())
    loop.run_until_complete(setup_webhook())

    t = threading.Thread(target=loop.run_forever, daemon=True)
    t.start()

    start_birthday_scheduler()

    logger.info(f"✅ Flask server port {PORT} da ishga tushdi!")
    flask_app.run(host="0.0.0.0", port=PORT)

if __name__ == "__main__":
    run()
